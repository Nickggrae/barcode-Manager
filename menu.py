# Nicholas Wharton
# Secure Password Manager
# Menu Controller
# 6/4/2024

import tkinter as tk
from barcode import scanSheet
from fileOperations import appendNewItem
from fileOperations import deleteItem
from fileOperations import copyInit
import openpyxl
import pandas
import os
from pygame import mixer

fontSize = 15

#menu for dealing with a sheet that already exists
class GetFilename(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller #controller of the window
        self.filename = tk.StringVar() #the filename of the book holding the scanned records

    #used when the page is loaded to refresh the displayed information
    def refresh(self):
        #destory the previous window information
        for widget in self.winfo_children():
            widget.destroy()

        prevFilename = ''

        #if the saved file has been created get the previously input filename if not create the saved file for next run
        if os.path.isfile('saved.xlsx'):
            book = openpyxl.load_workbook('saved.xlsx')
            sheet = book.active
            
            prevFilename = str(sheet.cell(row=1, column=1).value)
            self.filename.set(prevFilename)

            book.close()
        else:
            book = openpyxl.Workbook()
            book.save('saved.xlsx')
            book.close()

        tk.Label(self, text="Please Enter The Amazon Sheet or Copied Working Sheet", font=("Helvetica", fontSize)).pack(anchor="center")
        tk.Label(self, text="-----------------------------------------", font=("Helvetica", fontSize)).pack(anchor="center")
        tk.Label(self, text="Enter Filename: ", font=("Helvetica", fontSize)).pack(anchor="center")
        tk.Entry(self, textvariable=self.filename, font=("Helvetica", fontSize), width=75).pack(anchor="center")
        tk.Button(self, text="Continue", font=("Helvetica", fontSize), command=self.submit).pack(anchor="center")
        tk.Button(self, text="Quit", font=("Helvetica", fontSize), command=self.quit).pack(anchor="center")
        tk.Label(self, text="-----------------------------------------", font=("Helvetica", fontSize)).pack(anchor="center")

    #Start the scanning process on the existing file to append more records
    def submit(self):
        filename = self.filename.get()
        
        #update the most recently used filename to autofill when the menu is ran next time
        book = openpyxl.load_workbook('saved.xlsx')
        sheet = book.active
        sheet.cell(row=1, column=1).value = filename
        book.save('saved.xlsx')
        book.close()
        
        if filename == '': #no filename was input
            print("no filename was input")
        elif filename[0] == '2': #file is a amazon generated file
            newFile = copyInit(filename)
            self.controller.show_frame(SheetMenu, processedFilename=newFile)
        elif filename[0] == 'c': #file is a program generated file
            self.controller.show_frame(SheetMenu, processedFilename=filename)
        else:
            print("the file input is not supported")

    #end the program
    def quit(self):
       self.controller.quit()



#menu for dealing with a sheet that already exists
class SheetMenu(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller #controller of the window
        self.selectedRowIndex = tk.StringVar() #used to pass the users input to the processing functions
        self.filename = tk.StringVar() #the filename of the book holding the scanned records
        self.refreshes = 0
        self.currentBox = "None"
        self.lastScanned = ""
        self.currentItemList = []

    #used when the page is loaded to refresh the displayed information
    def refresh(self, **kwargs):
        #destory the previous window information
        for widget in self.winfo_children():
            widget.destroy()

        v = tk.Scrollbar(self)
        v.pack(side = "right", fill = "y")
        filename = ""

        #get the user input if possible
        if self.refreshes == 0:
            filename = str(kwargs.get("processedFilename"))
            self.filename.set(filename)
            self.refreshes = 1
        else:
            filename = self.filename.get()

        t = tk.Text(self, width = 15, height = 15, wrap = "none", yscrollcommand = v.set, font=("Courier", fontSize))

        t.insert("end", "-----------------------------------------\n")
        
        #open the book for reading the lines
        book = openpyxl.load_workbook(filename)
        sheet = book.active

        #Get the current number of boxes in the sheet.
        currentBoxNum = str(sheet.cell(row=1, column=2).value)

        #Get the number of item records
        ds = pandas.read_excel(filename)
        fileRows = ds.shape[0] + 1
        currentItemNum = fileRows - 2

        #list of current items records pulled from the file
        currentItemList = []

        #determine the length of the longest SKU in the list, add 4 ontop for proper formatting
        largestLength = 0
        i = 1
        while i <= currentItemNum:
            if len(str(sheet.cell(row=i + 2, column=1).value)) > largestLength:
                largestLength = len(str(sheet.cell(row=i + 2, column=1).value))
            i += 1

        largestLength += 4
        
        #labels for the displayed matrix
        labelStr = "Item"
        i = 4
        while i < largestLength:
            labelStr += " "

            i += 1

        labelStr += "|REQ|CUR|  |"
        t.insert("end", labelStr)

        i = 1
        while i <= int(currentBoxNum):
            t.insert("end", "B" + str(i) + "|")
            i += 1

        t.insert("end", "\n")


        #fill in the items and their current recorded values from the sheet
        i = 1
        while i <= currentItemNum:
            #first add the item SKU label
            t.insert("end", sheet.cell(row=i + 2, column=1).value.lower())
            
            #add space based on the size of the item SKU
            j = 0
            while j < (largestLength + 1 - len(sheet.cell(row=i + 2, column=1).value)):
                t.insert("end", " ")
                j += 1 

            #get the expected value string from the sheet
            t.insert("end", str(sheet.cell(row=i + 2, column=10).value))

            j = 0
            while j < (4 - len(str(sheet.cell(row=i + 2, column=10).value))):
                t.insert("end", " ")
                j += 1

            #calculate the sum of the item instances for the item object to get the current value
            currentInstancesCount = 0
            j = 0
            while j < int(currentBoxNum):
                if str(sheet.cell(row=i + 2, column= j + 13).value) != 'None':
                    currentInstancesCount += int(sheet.cell(row=i + 2, column= j + 13).value)
                j += 1

            #insert the current instances count into the presented string
            t.insert("end", str(currentInstancesCount))

            j = 0
            while j < (4 - len(str(currentInstancesCount))):
                t.insert("end", " ")
                j += 1

            #add the spacing to properly format the box matrix with its labels
            t.insert("end", "   ")

            #add each of the matrix values to the string
            j = 0
            while j < int(currentBoxNum):
                currentCellVal = str(sheet.cell(row=i + 2, column=j + 13).value)
                if (currentCellVal == "None"):
                    t.insert("end", "0  ")
                else:
                    t.insert("end", currentCellVal)
                    #add spacing based on the value of the cell (to deal with multi digit values messing up the spacing)
                    z = 0
                    while z < (3 - len(currentCellVal)):
                        t.insert("end", " ")
                        z += 1
                j += 1

            t.insert("end", "\n\n")

            i += 1

        t.pack(side="left", fill="both", expand=True)
        v.config(command=t.yview)

        book.close()

        #set currentItemList of window object so i can be accessed by the delete function
        self.currentItemList = currentItemList

        tk.Label(self, text="-----------------------------------------", font=("Helvetica", fontSize)).pack(anchor="center")
        tk.Button(self, text="Delete", font=("Helvetica", fontSize), command=self.delete).pack(anchor="center")
        tk.Button(self, text="Add More", font=("Helvetica", fontSize), command=self.append).pack(anchor="center")
        tk.Button(self, text="Back", font=("Helvetica", fontSize), command=self.back).pack(anchor="center")
        tk.Button(self, text="Quit", font=("Helvetica", fontSize), command=self.quit).pack(anchor="center")
        tk.Label(self, text="-----------------------------------------", font=("Helvetica", fontSize)).pack(anchor="center")
        tk.Label(self, text="Current Box Count: " + currentBoxNum, font=("Helvetica", fontSize)).pack(anchor="center")
        tk.Button(self, text="Add Box", font=("Helvetica", fontSize), command=self.addBox).pack(anchor="center")
        tk.Label(self, text="Current Box:", font=("Helvetica", fontSize)).pack(anchor="center")
        tk.Label(self, text=self.currentBox, font=("Helvetica", fontSize)).pack(anchor="center")
        tk.Label(self, text="Last Scanned:", font=("Helvetica", fontSize)).pack(anchor="center")
        tk.Label(self, text=self.lastScanned, font=("Helvetica", fontSize)).pack(anchor="center")
        tk.Label(self, text="\n\nPress Add More to enter *Scanning Mode*\n", font=("Helvetica", fontSize)).pack(anchor="center")
        tk.Label(self, text="Must scan a Box before scanning Items", font=("Helvetica", fontSize)).pack(anchor="center")
        tk.Label(self, text="To leave *Scanning Mode* Press '/'\n", font=("Helvetica", fontSize)).pack(anchor="center")
        tk.Label(self, text="If the window becomes unrespnsive click", font=("Helvetica", fontSize)).pack(anchor="center")
        tk.Label(self, text="any other window and press '/' to kill", font=("Helvetica", fontSize)).pack(anchor="center")
        tk.Label(self, text="the scanning mode process\n", font=("Helvetica", fontSize)).pack(anchor="center")
        tk.Label(self, text="To end the program press the quit button", font=("Helvetica", fontSize)).pack(anchor="center")

    #delete a record from the sheet based on user input and refresh the page
    def delete(self):
        filename = self.filename.get()
        
        currentBox = ""
        currentItem = ""

        #if a box has already been scanned set it to the current box so it doesnt have to be rescanned
        if self.currentBox != "None":
            currentBox = self.currentBox

        while True:
            #make the window unresponsive to clicks
            self.grab_set()
            item = str(scanSheet()) #get the next scanned 
            self.grab_release()

            if item == "'/'" or item == "": #make sure it was a barcode
                break
            else:
                if item[0] == 'B':
                    mixer.init()
                    mixer.music.load('bauDong.mp3')
                    mixer.music.play()
                    currentBox = item
                    self.currentBox = currentBox
                    self.lastScanned = item
                else:
                    currentItem = item
                    if currentBox == "":
                        print("No Box Associated with Item")
                    else:
                        print("currentBox " + currentBox + ", currentItem: " + currentItem)
                        itemFound = deleteItem(filename, currentBox, currentItem)
                        if itemFound == True:
                            mixer.init()
                            mixer.music.load('smallTellRing.mp3')
                            mixer.music.play()
                            self.lastScanned = item
                        else:
                            mixer.init()
                            mixer.music.load('fogHorn.mp3')
                            mixer.music.play()

                        currentItem = ""
                self.refresh()
                self.update()
        
        self.refresh()

    #Start the scanning process on the existing file to append more records
    def append(self):
        filename = self.filename.get()
        
        currentBox = ""
        currentItem = ""

        #if a box has already been scanned set it to the current box so it doesnt have to be rescanned
        if self.currentBox != "None":
            currentBox = self.currentBox

        while True:
            #make the window unresponsive to clicks
            self.grab_set()
            item = str(scanSheet()) #get the next scanned 
            self.grab_release()

            if item == "'/'" or item == "": #make sure it was a barcode
                break
            else:
                if item[0] == 'B':
                    mixer.init()
                    mixer.music.load('bauDong.mp3')
                    mixer.music.play()
                    currentBox = item
                    self.currentBox = currentBox
                    self.lastScanned = item
                else:
                    currentItem = item
                    if currentBox == "":
                        print("No Box Associated with Item")
                    else:
                        print("currentBox " + currentBox + ", currentItem: " + currentItem)
                        itemFound = appendNewItem(filename, currentBox, currentItem)
                        if itemFound == True:
                            mixer.init()
                            mixer.music.load('smallTellRing.mp3')
                            mixer.music.play()
                            self.lastScanned = item
                        else:
                            mixer.init()
                            mixer.music.load('fogHorn.mp3')
                            mixer.music.play()
                        currentItem = ""
                self.refresh()
                self.update()
        
        self.refresh()

    #Increments the total boxes value in the working sheet
    def addBox(self):
        filename = self.filename.get()
        
        #open the book for reading the lines
        book = openpyxl.load_workbook(filename)
        sheet = book.active

        sheet.cell(row=1, column=2).value = int(sheet.cell(row=1, column=2).value) + 1
        
        book.save(filename)
        book.close()
        
        self.refresh()

    def back(self):
        self.controller.show_frame(GetFilename)

    #end the program
    def quit(self):
        os.system(f'start excel {self.filename.get()}')
        self.controller.quit()


#master class for the application
class Application(tk.Tk):
    #initalize the application window and set the basic setting
    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)
        container = tk.Frame(self) #controller of the window
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}
        #initalize which frames exist
        for F in (SheetMenu, GetFilename):
            frame = F(container, self)
            self.frames[F] = frame
            frame.grid(row=0, column=0, sticky="nsew")

        #the first frame to open on startup
        self.show_frame(GetFilename)

    #function for switching between frames
    def show_frame(self, cont, **kwargs):
        frame = self.frames[cont]
        frame.refresh(**kwargs)
        frame.tkraise()



if __name__ == "__main__":
    app = Application()
    app.geometry("1200x900")
    app.mainloop()