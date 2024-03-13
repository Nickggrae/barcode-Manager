#Nicholas Wharton
#Barcode Scanner Item Managment
#used to build the invoice from the processed items sheet
#3/11/2024

import openpyxl
from openpyxl.styles import PatternFill
import pandas
from datetime import datetime


#driver function for building a invoice from the processed item information
def buildInvoice(inputFile, boxFile):
    
    uniqueItems, itemsBoxes, boxes = getItemInfo(inputFile, boxFile) #get the item information needed to build the invoice from the input file

    outputFile = generateInvoice(uniqueItems, itemsBoxes, inputFile, boxes) #generate the invoice with the collected item information
    
    return outputFile


#Get the item information needed to build the invoice from the input file
def getItemInfo(inputFile, boxFile):
    uniqueItems = [] #array of unique items that were scanned
    itemsBoxes = [] #array of arrays of each related to the uniqueItems index showing which boxes contain the item
    boxes = [] #array of box names pulled form the boxFile

    #how many lines are there in the boxFile
    ds = pandas.read_excel(boxFile)
    fileRows = ds.shape[0] + 1

    #open the boxfile
    book = openpyxl.load_workbook(boxFile)
    sheet = book.active

    #store all the box names into a list of strings
    i = 1
    while i <= fileRows:
        boxes.append(sheet.cell(row=i, column=1).value)
        i += 1

    book.close()

    #how many lines are there in the boxFile
    ds = pandas.read_excel(inputFile)
    fileRows = ds.shape[0] + 1

    #open the boxfile
    book = openpyxl.load_workbook(inputFile)
    sheet = book.active

    #store all the box names into a list of strings
    i = 1
    while i <= fileRows:
        #grab the first items FNSKU and its corresponding box
        currItem = sheet.cell(row=i, column=3).value
        currItemBox = sheet.cell(row=i, column=1).value

        print("currItem: " + currItem + "    currItemBox: " + currItemBox)

        #check if this item already has a record in the unique items array
        isDistinct = True
        j = 0
        while j < len(uniqueItems):
            if currItem == uniqueItems[j]:
                isDistinct = False
            j += 1

        #if the item is the first instance of the item in the file
        if isDistinct == True:
            uniqueItems.append(currItem) #append the new item to the list of unique items
            itemsBoxes.append([]) #start a new box array for the item
            itemsBoxes[len(itemsBoxes) - 1].append(currItemBox) #add the items box number as the first instance in the item's item box array
        else:
            j = 0
            while j < len(uniqueItems): #find where the items record is stored
                if currItem == uniqueItems[j]:
                    itemsBoxes[j].append(currItemBox) #add the items box to its current box record array
                j += 1

        i += 1

    book.close()

    #print the uniqueItems and itemsBoxes array info to console for testing
    i = 0
    for item in uniqueItems:
        print(str(i) + ": " + item)
        i += 1

    i = 0
    for box in itemsBoxes:
        print(str(i) + ": ")
        print(box)
        i += 1

    return uniqueItems, itemsBoxes, boxes



def generateInvoice(uniqueItems, itemsBoxes, inputFile, boxes):
    #how many lines are there in the boxFile
    ds = pandas.read_excel(inputFile)
    fileRows = ds.shape[0] + 1

    #create a new sheet for the invoice
    book = openpyxl.Workbook()
    sheet = book.active

    #add the labels and basic information to the invoice sheet
    #sheet.cell(row=1, column=1, value="Provide the box details for this pack group below. See the instructions sheet if you have more questions")
    sheet.cell(row=2, column=1, value="Pack Group: 1")
    sheet.cell(row=3, column=1, value="Total SKUs: " + str(len(uniqueItems)) + " (" + str(fileRows) + " units)")
    cell = sheet.cell(row=2, column=1)
    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    cell = sheet.cell(row=3, column=1)
    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


    #Add the item inforamtion labels
    sheet.cell(row=5, column=1, value="Product Title")
    sheet.cell(row=5, column=2, value="FNSKU")
    sheet.cell(row=5, column=3, value="SKU")
    sheet.cell(row=5, column=4, value="ID")

    #color labeled and empty cells
    for b in range(1, 5):
        cell = sheet.cell(row=5, column=b)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    #add the label for each box in the box file
    i = 1
    for box in boxes:
        sheet.cell(row=5, column=6+i, value=box)
        cell = sheet.cell(row=5, column=6+i)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        i += 1

    sheet.cell(row=3, column=5, value="Total Box Count:")
    sheet.cell(row=3, column=6, value=str(len(boxes)))

    cell = sheet.cell(row=3, column=5)
    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    cell = sheet.cell(row=3, column=6)
    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    #open the inputFile to get each items information
    inBook = openpyxl.load_workbook(inputFile)
    inSheet = inBook.active

    #add the information for each unique item into the sheet
    i = 0
    while i < len(uniqueItems):
        j = 1
        #for each item in the input file
        while j <= fileRows:
            #if the item is one of the unique items copy its data into the invoice 
            if uniqueItems[i] == inSheet.cell(row=j, column=3).value:
                print(inSheet.cell(row=j, column=2).value + "    " + inSheet.cell(row=j, column=3).value)

                sheet.cell(row=i + 6, column=1, value=inSheet.cell(row=j, column=2).value)
                sheet.cell(row=i + 6, column=2, value=inSheet.cell(row=j, column=3).value)
                sheet.cell(row=i + 6, column=3, value=inSheet.cell(row=j, column=4).value)
                sheet.cell(row=i + 6, column=4, value=inSheet.cell(row=j, column=5).value)
                break
            j += 1

        #fill in the box item association matrix
        for box in itemsBoxes[i]:
            characters = [char for char in box] #break the box string into an array of characters
            boxNumber = int(characters[9]) #get the number value from the box string

            #determine if a value exists for the relation between the item's instance # and the box
            if sheet.cell(row=i + 6, column=boxNumber + 7).value is None:
                #initalize the instance # assocaited with the box to 1
                sheet.cell(row=i + 6, column=boxNumber + 7, value="1")
                cell = sheet.cell(row=i + 6, column=boxNumber + 7)
                cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
            else:
                #increment the current instance # assocaited with the box by 1
                sheet.cell(row=i + 6, column=boxNumber + 7, value=str(int(sheet.cell(row=i + 6, column=boxNumber + 7).value) + 1))  
        
        i += 1 #increment to next unqiue item index
    
    
    #color empty cells
    for b in range(1, 7 + len(boxes)):
        cell = sheet.cell(row=4, column=b)
        cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        cell = sheet.cell(row=i + 6, column=b)
        cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        cell = sheet.cell(row=i + 12, column=b)
        cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    i += 7 #skip one more cell and add the hardcoded offset like above but to the variable to give space between box matrix and the box dimension and weight fields

    #color empty cells
    for b in range(4, i + 5):
        cell = sheet.cell(row=b, column=5)
        cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    for b in range(4, i + 6):
        cell = sheet.cell(row=b, column=7 + len(boxes))
        cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    #more labels
    sheet.cell(row=i, column=6, value="Name of Box")
    sheet.cell(row=i + 1, column=6, value="Box Weight (lbs):")
    sheet.cell(row=i + 2, column=6, value="Box Width (inch):")
    sheet.cell(row=i + 3, column=6, value="Box Length (inch):")
    sheet.cell(row=i + 4, column=6, value="Box Height (inch):")

    #color empty cells
    for b in range(0, 5):
        cell = sheet.cell(row=i + b, column=6)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    for b in range(5, 6 + len(uniqueItems)):
        cell = sheet.cell(row=b, column=6)
        cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    #Label for each box
    c = 1
    while c <= len(boxes):
        sheet.cell(row=i, column=6 + c, value="P1 - B" + str(c - 1))
        cell = sheet.cell(row=i, column=6 + c)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        c += 1

    # c = 1
    # while c <= len(boxes)




    #set the colum width for each cell to allow the user to see all of the information when they open the sheet
    columnMaxLen = {}
    for row in sheet.iter_rows(values_only=True):
        for i, cellText in enumerate(row, start=1):
            if i not in columnMaxLen:
                columnMaxLen[i] = len(str(cellText)) + 10
            else:
                columnMaxLen[i] = max(columnMaxLen[i], len(str(cellText)) + 5)

    # Set column width based on the maximum length of text
    for column, maxLen in columnMaxLen.items():
        sheet.column_dimensions[sheet.cell(row=1, column=column).column_letter].width = maxLen

    


    inBook.close()
    #sheet.cell(row=i, column=j, value=part)


    #generate a name for the invoice and save the file
    now = datetime.now()
    outputFilename = 'invoice' + str(now.strftime("%m-%d-%H-%M")) + '.xlsx'
    print(outputFilename)

    book.save(outputFilename)
    book.close()
    return outputFilename