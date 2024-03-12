import openpyxl
from datetime import datetime

#used to convert the invoice sheets into the list of item records so they can be processed in the program
def invoiceToProcessed(invoiceFilename):
    print("Processed To Invoice Conversion Initiated.")

    #open the invoice file
    invoiceBook = openpyxl.load_workbook(invoiceFilename)
    invoiceSheet = invoiceBook.active

    boxCount = int(invoiceSheet.cell(row=3, column=6).value)

    #get the total SKUs cell and get the number of unique items to know how many 
    totalSKUs = invoiceSheet.cell(row=3, column=1).value
    totalSKUs = totalSKUs.split()
    uniqueItems = int(totalSKUs[2])

    #create a new sheet for the processed file
    book = openpyxl.Workbook()
    sheet = book.active

    instanceCount = 1
    i = 0
    while i < uniqueItems:
        j = 0
        while j < boxCount:
            tempCellVal = invoiceSheet.cell(row=i + 6, column=7 + j).value
            if (tempCellVal != None):
                z = 0
                while z < int(tempCellVal):
                    sheet.cell(row=instanceCount, column=1, value=invoiceSheet.cell(row=5, column=7 + j).value)
                    sheet.cell(row=instanceCount, column=2, value=invoiceSheet.cell(row=i + 6, column=1).value)
                    sheet.cell(row=instanceCount, column=3, value=invoiceSheet.cell(row=i + 6, column=2).value)
                    sheet.cell(row=instanceCount, column=4, value=invoiceSheet.cell(row=i + 6, column=3).value)
                    sheet.cell(row=instanceCount, column=5, value=invoiceSheet.cell(row=i + 6, column=4).value)
                    instanceCount += 1
                    z += 1
            j += 1
        i += 1

    #generate a name for the invoice and save the file
    now = datetime.now()
    outputFilename = 'changed' + str(now.strftime("%m-%d-%H-%M")) + '.xlsx'
    print(outputFilename)

    book.save(outputFilename)

    return outputFilename