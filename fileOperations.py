#Nicholas Wharton
#Barcode Scanner Item Managment
#file operations to manage scanned barcode data
#6/4/2024

import openpyxl
import pandas
from datetime import datetime


#takes the amazon file and configures the working sheet so it has all of the modularity features 
#that the amazon sheet has, and it labeled in the same way so when you are done the item matrix 
#can be copied over to the amazon file.
def copyInit(azFile):

    #Open the amazon input file
    azBook = openpyxl.load_workbook(azFile)
    azSheet = azBook.active

    #Create the new working file
    book = openpyxl.Workbook()
    sheet = book.active

    #grab box count and calculate the amount of columns in the sheet
    boxCount = int(azSheet.cell(row=3, column=13).value)
    maxColums = 12 + boxCount

    #grab the item instance number and calculate the amount of rows in the sheet
    totalSku = str(azSheet.cell(row=3, column=1).value).split()
    maxRows = int(totalSku[2]) + 3

    #Put the box value in the working sheet
    sheet.cell(row=1, column=1).value = "Total Box(s):"
    sheet.cell(row=1, column=2).value = boxCount

    #Put the Total Item Counter and Label
    sheet.cell(row=1, column=3).value = "Total Item(s):"
    sheet.cell(row=1, column=4).value = "=COUNTA(A:A)-2"

    #Put the Total Item Instance Counter and Label
    sheet.cell(row=1, column=5).value = "Total Item Instances:"
    sheet.cell(row=1, column=6).value = "=SUM(K:K)"

    #copy the colum labels 
    i = 1
    while i < 12:
        sheet.cell(row=2, column=i).value = azSheet.cell(row=5, column=i).value
        i += 1

    #copy the unique item instance data
    i = 3
    while i < maxRows:
        j = 1
        while j < maxColums:
            sheet.cell(row=i, column=j).value = azSheet.cell(row=i + 3, column=j).value
            j += 1
        
        i += 1

    #setup box labels to appear based on the input box value so it can be manually changed
    #23 for 10 box maximum
    i = 1
    while i < 10:
        sheet.cell(row=2, column=i + 12).value = '=IF(B1>=' + str(i) + ', "Box ' + str(i) + '","")'
        i += 1

    #Add the boxed quantity formula so it dynamically counts the amount of item instances for each item record
    print("Max Rows: " + str(maxRows))
    i = 3
    while i < maxRows:
        sheet.cell(row=i, column=11).value = '=SUM(M' + str(i) + ':OFFSET(M' + str(i) + ', 0, B1-1))'
        i += 1

    #shorten the header for expected and boxed quantity purly for looks
    sheet.cell(row=2, column=10).value = 'Expected'
    sheet.cell(row=2, column=11).value = 'Boxed'

    #close the amazon sheet
    azBook.close()

    #resize all of the cells to make all of the cell values visable
    i = 1
    for column in sheet.columns:
        if i == 1 or i == 3 or i == 4 or i == 5 or i == 10:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width
        i += 1

    #name the working sheet, save it and close it.
    now = datetime.now()
    outputFilename = 'copiedSheet' + str(now.strftime("%m-%d-%H-%M")) + '.xlsx'

    book.save(outputFilename)
    book.close()

    #return the filename so it can be accessed by the other funtions 
    return outputFilename




#for each new item given the working file and the box its supposed to go into update the item instance matrix in the given file.
def appendNewItem(filename, currentBox, currentItem):
    #determine how many records are in the file
    ds = pandas.read_excel(filename)
    fileRows = ds.shape[0] + 1
    currentItemNum = fileRows - 2
    
    book = openpyxl.load_workbook(filename)
    sheet = book.active
    
    currentBoxNum = int(currentBox[9])

    i = 1
    while i < currentItemNum:
        indexedItem = str(sheet.cell(row=i + 2, column=5).value)
        if indexedItem == currentItem:
            if sheet.cell(row=i + 2, column=12 + currentBoxNum).value == None:
                sheet.cell(row=i + 2, column=12 + currentBoxNum).value = 1
            else:
                sheet.cell(row=i + 2, column=12 + currentBoxNum).value = int(sheet.cell(row=i + 2, column=12 + currentBoxNum).value) + 1

        i += 1

    book.save(filename)
    book.close()


#for each box-item pair scanned given the working file decrement the corresponding item instance matrix cell.
def deleteItem(filename, currentBox, currentItem):
    #determine how many records are in the file
    ds = pandas.read_excel(filename)
    fileRows = ds.shape[0] + 1
    currentItemNum = fileRows - 2
    
    book = openpyxl.load_workbook(filename)
    sheet = book.active
    
    currentBoxNum = int(currentBox[9])

    i = 1
    while i < currentItemNum:
        indexedItem = str(sheet.cell(row=i + 2, column=5).value)
        if indexedItem == currentItem:
            if sheet.cell(row=i + 2, column=12 + currentBoxNum).value != None:
                sheet.cell(row=i + 2, column=12 + currentBoxNum).value = int(sheet.cell(row=i + 2, column=12 + currentBoxNum).value) - 1

        i += 1

    book.save(filename)
    book.close()


#appendNewItem("copiedSheet05-22-15-19.xlsx", "BOX0000001", "PPB-Kin-Hick-BBQ-2pk-3Lid")
#copyInit("2024-03-13_02-17-15_pack-group-1_99e4d13a-e977-4009-adb0-5ff8d8457be5.xlsx")