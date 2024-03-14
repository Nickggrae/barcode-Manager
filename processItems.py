#Nicholas Wharton
#Barcode Scanner Item Managment
#Process scanned info
#3/11/2024

import openpyxl
import pandas
from datetime import datetime


def process (inputFile, currentBox, currentItem, boxFile, itemsFile):
    boxes = [] #list of all the boxes from the boxFile
    tempCompData = [] #holds a item record with its pulled item data

    print("Starting Processing of Scanned Items...")
    
    #how many lines are there in the boxFile
    ds = pandas.read_excel(boxFile)
    fileRows = ds.shape[0] + 1

    #open the boxfile
    book = openpyxl.load_workbook(boxFile)
    sheet = book.active

    #store all the box names into a list of strings
    i = 1
    while i <= fileRows:
        boxes.append(sheet.cell(row=i, column=1).value)
        i += 1

    book.close()

    #--------------------------------------------------------------------------------------------

    #how many lines are there in the item file
    ds = pandas.read_excel(itemsFile)
    fileRows = ds.shape[0] + 1

    #open the item records file
    book = openpyxl.load_workbook(itemsFile)
    sheet = book.active
        

    #search for the item based on its FNSKU
    i = 1 #iterate through each row
    while i <= fileRows:
        if currentItem == sheet.cell(row=i, column=2).value:
            tempCompData.append(currentBox)
            tempCompData.append(sheet.cell(row=i, column=1).value)
            tempCompData.append(sheet.cell(row=i, column=2).value)
            tempCompData.append(sheet.cell(row=i, column=3).value)
            tempCompData.append(sheet.cell(row=i, column=4).value)
        i += 1
    
    book.close()
    
    #if the item has no data in the record sheet then still add it to the sheet but showing no data was found
    if len(tempCompData) == 0:
        print("Item " + currentItem + " Has No Info In Item Record Field")
        tempCompData.append(currentBox)
        tempCompData.append("N/A")
        tempCompData.append(currentItem)



    #add the data to the sheet
    ds = pandas.read_excel(inputFile)
    fileRows = ds.shape[0] + 1

    #if a input file is passed append the scanned records to the existing file
    book = openpyxl.load_workbook(inputFile)
    sheet = book.active

    print(tempCompData)

    j = 1
    for part in tempCompData:
        print(part)
        sheet.cell(row=fileRows+1, column=j, value=part)
        j += 1

    book.save(inputFile)
    book.close()

    return True